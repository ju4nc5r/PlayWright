import openpyxl
import csv
import os
import sys
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.borders import BORDER_THIN, BORDER_NONE
from openpyxl.styles.colors import (BLACK, RED, DARKRED, BLUE, DARKBLUE,
                                    GREEN, DARKGREEN, YELLOW, DARKYELLOW,
                                    WHITE)
from SeleniumFramework.src.utils.constants.excel_file_constants import (
    C_BLACK, C_RED, C_DARKRED, C_BLUE, C_DARKBLUE, C_GREEN, C_DARKGREEN,
    C_YELLOW, C_DARKYELLOW, C_WHITE, SOLID, TOP, LEFT
)
from SeleniumFramework.src.utils.settings import hb_user_excel, hb_user_login

python_version = sys.version_info[0]  # Get the python version
if python_version == 2:
    from openpyxl.cell.cell import Cell
elif python_version == 3:
    from openpyxl.cell.cell import Cell
    from openpyxl.cell.cell import TYPE_FORMULA


class excel_file():
    def __init__(self, path, sheet, data_only=False, read_only=False):
        """
        Start a driver for excel.
        :param path: String. Path of file.
        :param sheet: String. Name of sheet.
        :param data_only: Boolean. If read formulas or results from cells.
        :param read_only: Boolean. Protect the file.
        """
        self.__path = path
        self.__sheet_name = sheet
        # Si no existe el archivo, se crea
        if not os.path.exists(self.__path):
            self.__create_workbook()
        self.__xl = openpyxl.load_workbook(self.__path, data_only=data_only,
                                           read_only=read_only)
        self.__sheet = self.__xl[self.__sheet_name]
        self.__head = self.read_row(1)

    def __create_workbook(self):
        """
        Create a new workbook and change name from active sheet.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self.__sheet_name
        wb.save(self.__path)

    def __read_cell_value(self, cell):
        """
        Method to read a cell value. If the cell is a formula, read the result
        :param cell: Cell. Cell from the sheet
        """
        if self.__is_formula(cell):
            xl_aux = openpyxl.load_workbook(self.__path, data_only=True)
            sheet_aux = xl_aux[self.__sheet_name]
            value = sheet_aux[cell.coordinate].value
            return value
        value = cell.value
        return value

    def read_cells(self, cell_range):
        """
        Method to read a range of cells.
        :param cell_range: String. Range of cells.
        :return: List(List). Array of Arrays with value of cells.
        """
        aux = []
        cells = self.__get_cells(cell_range)
        if len(cells) == 1:
            cell = cells[0]
            return self.__read_cell_value(cell)
        for cell in cells:
            value = self.__read_cell_value(cell)
            aux.append(value)
        return aux

    def read_row(self, row):
        """
        Method to read a row in the sheet.
        :param row: Integer. Number of row to read.
        :return: List. List with value of cells.
        """
        row = self.__sheet[row]
        aux = []
        for cell in row:
            value = self.__read_cell_value(cell)
            aux.append(value)
        return aux

    def read_column(self, column):
        """
        Method to read a column from the excel file.
        :param column: Integer. Number of the column to read.
        :return: List. List with values of column
        """
        col = self.get_column(column)
        aux = []
        for cell in col:
            value = self.__read_cell_value(cell)
            aux.append(value)
        return aux

    def get_column(self, column):
        """
        Method to obtain the cell from a column.
        :param column: Integer. Number of the column to read.
        :return: List of cell from column
        """
        cols = self.get_cols()
        if python_version == 2:
            for _ in range(column):
                col = cols.next()
        elif python_version == 3:
            counter = 1
            for i in cols:
                col = i
                if column == counter:
                    break
                counter += 1
        return col

    def read_row_by_head(self, row):
        """
        Method to read a row in the sheet and, return a dict
        :param row: Integer. Number of row to read.
        :return: dict. Dict with value of cell by head
        """
        if self.__head is not None:
            row = self.read_row(row)
            d = dict()
            for i in range(len(self.__head)):
                d[self.__head[i]] = row[i]
        return d

    def write_row_by_head(self, row, data):
        """
        Method to write new values in a row by head
        :param row: Integer. Number of row to write
        :param data: dict. New data to update in the excel file
        """
        row_data = self.read_row_by_head(row)
        row_data.update(data)
        first_row = self.__sheet[1]
        first_row_value = self.__head
        for key, value in row_data.items():
            if key in first_row_value:
                if value is not None:
                    index = first_row_value.index(key)
                    col = first_row[index].column
                    self.write_cell_by_cell(col+str(row), value)
                    # self.write_cell_by_cell("A"+str(row), value)
            else:
                print("The value '{}' isn't in head".format(key))

    def __get_cells(self, cell_range):
        """
        Method to obtain all of cells in the range.
        :param cell_range: String. Range of cells.
        :return: List. List of Cell
        """
        string = cell_range.split(':')
        if len(string) == 1:
            return [self.__sheet[cell_range]]
        aux = [cell for row in self.__sheet[cell_range] for cell in row]
        return aux

    def __enter_value(self, cell, text):
        """
        Method to input a value in cell.
        :param cell: Cell. A cell in the sheet.
        :param text: String. Text to input.
        """
        print("Cell en enter_value "+ str(cell))
        print ("Texto en enter_value "+ str(text))
        cell.value = text
        self.__save()

    def write_cell_by_index(self, row, column, text):
        """
        Method to write a text in a cell. Row and column is passed as
        parameter.
        :param row: Integer. Row of cell. Start at 1.
        :param column: Integer. Column of cell. Start at 1.
        :param text: Strign. Text to input in cell.
        """
        cell = self.get_cell_by_index(row, column)
        self.__enter_value(cell, text)

    def write_cell_by_cell(self, cell_range, text):
        """
        Method to write in a range of cells.
        :param cell_range: Strign. Range of cells.
        :param text: Strign. Text to input in cells.
        """
        
        cells = self.__get_cells(cell_range)
        for cell in cells:
            self.__enter_value(cell, text)
        self.__save()

    def set_font(self, cell_range=None, cell=None, name=None, sz=None, b=None,
                 i=None, charset=None, u=None, strike=None, color=None,
                 scheme=None, family=None, size=None, bold=None, italic=None,
                 strikethrough=None, underline=None, vertAlign=None,
                 outline=None, shadow=None, condense=None, extend=None,
                 font=None):
        """
        Method to apply a font in a cell range.
        :param cells_range: String. Range of cells.
        :param cell: Cell. If you want to change to a specific cell.
        :param name: String. Name of font to use.
        :param strike: Boolean. If you want strikethrough.
        :param color: String. The color you want use.
        :param size: Integer. Size of letter in the cell.
        :param bold: Boolean. If you want use bold in the cells.
        :param italic: Boolean. If you want use italic in the cells.
        :param underline: String. If you want use underline. The option of
                          underline are in the Font class as UNDERLINE_
        :param font: Font. Font Class to use. Default is None
        """
        if color is not None:
            color = self.get_color(color)
        if font is None:
            font = Font(name, sz, b, i, charset, u, strike, color, scheme,
                        family, size, bold, italic, strikethrough, underline,
                        vertAlign, outline, shadow, condense, extend)
        if cell is not None:
            cells = [cell]
        elif cell_range is not None:
            cells = self.__get_cells(cell_range)
        for c in cells:
            c.font = font

    def __set_font(self, cell, font):
        """
        Method to set a font in a cell.
        :param cell: Cell. Cell to apply the font.
        :param font: Font. Font to use.
        """
        f = Font(font.name, font.sz, font.b, font.i, font.charset, font.u,
                 font.strike, font.color, font.scheme, font.family, font.size,
                 font.bold, font.italic, font.strikethrough, font.underline,
                 font.vertAlign, font.outline, font.shadow, font.condense,
                 font.extend)
        cell.font = f

    def set_inner_border(self, cell_range, style=BORDER_THIN, color=C_BLACK):
        """
        Method to apply inner border in a range of cells.
        :param cell_range: String. Range of cells.
        :param style: String. Style of border to use.
        :param color: String. Color to use.
        """
        color = self.get_color(color)
        side = Side(border_style=style, color=color)
        border = Border(left=side, right=side, top=side, bottom=side)
        for cell in self.__get_cells(cell_range):
            cell.border = border
        self.__save()

    def set_border(self, cell_range, style=BORDER_THIN, color=C_BLACK):
        """
        Method to apply boder in outside of range of cells.
        :param cell_range: String. Range of cells.
        :param style: String. Style of border to use.
        :param color: String. Color to use.
        """
        rows = self.__sheet[cell_range]
        color = self.get_color(color)
        side = Side(border_style=style, color=color)

        rows = list(rows)
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if (pos_x == 0 or pos_x == max_x or
                        pos_y == 0 or pos_y == max_y):
                    cell.border = border
        self.__save()

    def __set_border(self, cell, border):
        """
        Method to apply a border in a cell.
        :param cell: Cell. Cell to apply the border.
        :param border: Border. Border to apply.
        """
        b = Border(border.left, border.right, border.top, border.bottom,
                   border.diagonal, border.diagonal_direction,
                   border.vertical, border.horizontal, border.diagonalUp,
                   border.diagonalDown, border.outline, border.start,
                   border.end)
        cell.border = b

    def reset_border(self, cell_range):
        """
        Method to delete border form a range of cells.
        :param cell_range: String. Range of cells.
        """
        side = Side(border_style=BORDER_NONE)
        border = Border(left=side, right=side, top=side, bottom=side)
        for cell in self.__get_cells(cell_range):
            cell.border = border
        self.__save()

    def set_fill(self, cell_range=None, cell=None, fill_type=SOLID,
                 start_color=None, end_color=None):
        """
        Method to set fill in range of cells.
        :param cell_range: String. Range of cells.
        :param cell: Cell. If you want change pattern fill only a cell.
        """
        fill = PatternFill(fill_type=fill_type, start_color=start_color,
                           end_color=end_color)
        if cell is not None:
            cells = [cell]
        elif cell_range is not None:
            cells = self.__get_cells(cell_range)
        for c in cells:
            c.fill = fill
        self.__save()

    def set_cell_fill(self, color):
        """
        Method to apply a pattern fill in a cell.
        :param color: String. Color to use.
        :return fill: PatternFill. Fill to use.
        """
        color = self.get_color(color)
        fill = PatternFill(
            start_color=color, end_color=color, fill_type=SOLID)
        return fill

    def __reset_cell_fill(self, cell):
        """
        Method to delete pattern fill in a cell.
        """
        color = self.get_color(C_WHITE)
        fill = PatternFill(
            start_color=color, end_color=color, fill_type=SOLID)
        cell.fill = fill

    def __set_fill(self, cell, fill):
        f = PatternFill(fill.patternType, fill.fgColor, fill.bgColor,
                        fill.fill_type, fill.start_color, fill.end_color)
        cell.fill = f

    def set_alignment(self, cell_range=None, cell=None, vertical=TOP,
                      horizontal=LEFT, textRotation=0):
        """
        Method to set a alignment in a range of cells.
        :param vertical: String. Value of vertical sense. The options are in
                         vertical_alignments in Alignment.
        :param horizontal: Strign. Value of horizontal sense. The options
                           are in horizontal_alignments in Alignment.
        :param textRotation: Integer. Value to ratate text in the cell.
        """
        alig = Alignment(horizontal=horizontal, vertical=vertical,
                         textRotation=textRotation)
        if cell is not None:
            cells = [cell]
        elif cell_range is not None:
            cells = self.__get_cells(cell_range)
        for c in cells:
            c.alignment = alig
        self.__save()

    def __set_alignment(self, cell, alig):
        a = Alignment(alig.horizontal, alig.vertical, alig.textRotation,
                      alig.wrapText, alig.shrinkToFit, alig.indent,
                      alig.relativeIndent, alig.justifyLastLine,
                      alig.readingOrder, alig.text_rotation, alig.wrap_text,
                      alig.shrink_to_fit)
        cell.alignment = a

    def get_color(self, color):
        """
        Method to obtain hexa value of a color
        :param color: String. Name of color to use.
        :return: String. Hexa value of a color.
        """
        if color.lower() == C_BLACK:
            return BLACK
        elif color.lower() == C_RED:
            return RED
        elif color.lower() == C_DARKRED:
            return DARKRED
        elif color.lower() == C_BLUE:
            return BLUE
        elif color.lower() == C_DARKBLUE:
            return DARKBLUE
        elif color.lower() == C_GREEN:
            return GREEN
        elif color.lower() == C_DARKGREEN:
            return DARKGREEN
        elif color.lower() == C_YELLOW:
            return YELLOW
        elif color.lower() == C_DARKYELLOW:
            return DARKYELLOW
        elif color.lower() == C_WHITE:
            return WHITE
        return color

    def clear_cell(self, cell):
        """
        Method to delete value from a cell.
        :param cell: Cell. Cell to delete value.
        """
        cell.value = None

    def clear_cell_by_range(self, cell_range):
        """
        Method to delete value from a range of cells.
        :param cell_range: String. Range of cells.
        """
        cells = self.__get_cells(cell_range)
        for cell in cells:
            self.clear_cell(cell)

    def clear_column(self, column):
        """
        Method to delete value from cells in the column.
        :param column: Integer. Column to delete value.
        """
        col = self.get_column(column)
        for cell in col:
            self.clear_cell(cell)
            self.__reset_cell_fill(cell)
        self.__save()

    def clear_row(self, row):
        """
        Method to delete value from cells in the row.
        :param row: Integer. Row to delete value.
        """
        ro = self.__sheet[row]
        for cell in ro:
            self.clear_cell(cell)
            self.__reset_cell_fill(cell)
        self.__save()

    def delete_cols(self, column):
        """
        Method to delete a column from the sheet.
        :param column: Integer. Column to delete.
        """
        self.__sheet.delete_cols(column)
        self.__save()

    def delete_rows(self, row):
        """
        Method to delete a row from the sheet.
        :param row: Integer. Row to delete.
        """
        self.__sheet.delete_rows(row)
        self.__save()

    def create_sheet(self, name):
        """
        Method to create a new sheet in workbook.
        :param name: String. Name of new sheet.
        """
        self.__xl.create_sheet(name)
        self.__save()

    def delete_sheet(self, name):
        """
        Method to delete a sheet from the workbook.
        :param name: String. Name of sheet to delete.
        """
        self.__xl.remove(self.__xl[name])
        self.change_sheet(0)
        self.__save()

    def change_sheet(self, index):
        """
        Method to change active sheet
        :param index: Integer. Index of new active sheet
        """
        self.__xl.active = index
        self.__sheet = self.__xl.active

    def merge_cells(self, cell_range):
        """
        Method to merge cells in range of cells.
        :param cell_range: String. Range of cells to merge.
        """
        self.__sheet.merge_cells(cell_range)
        self.__save()

    def unmerge_cells(self, cell_range):
        """
        Method to unmerge cells in range of cells.
        :param cell_range: String. Range of cells.
        """
        self.__sheet.unmerge_cells(cell_range)
        self.__save()

    def insert_image(self, cell, image_path):
        """
        Method to insert a image in workbook
        :param cell: Cell. Cell to insert the image
        :param image: Image. Image to insert
        """
        img = Image(image_path)
        self.__sheet.add_image(img, cell)
        self.__save()

    def copy_range(self, cell_range, paste_cell):
        """
        Method to copy a range of cell in other cell.
        :param cell_range: String. Range of cell to copy
        :param paste_cell: String. Cell to paste.
        """
        copy_cells = self.__sheet[cell_range]
        paste_cell = self.__sheet[paste_cell]
        paste_row = 0
        try:
            # Si se quiere copiar un rango de celdas
            for row in copy_cells:
                paste_column = 0
                for cell in row:
                    paste = paste_cell.offset(paste_row, paste_column)
                    self.__paste_cell(paste, cell)
                    paste_column += 1
                paste_row += 1
        except TypeError:
            # Si se quiere copiar solo una celda
            self.__paste_cell(paste_cell, copy_cells)
        finally:
            self.__save()

    def __paste_cell(self, paste_cell, copy_cell):
        """
        Copy the format of cell, and pasta in new cell
        :param paste_cell: Cell. Cell to paste
        :param copy_cell: Cell. Cell to copy
        """
        paste_cell.value = copy_cell.value
        self.__set_font(paste_cell, copy_cell.font)
        self.__set_fill(paste_cell, copy_cell.fill)
        self.__set_border(paste_cell, copy_cell.border)
        self.__set_alignment(paste_cell, copy_cell.alignment)

    def convert_csv(self, csv_path):
        """
        Convert the excel to csv
        :param csv_path: String. Path of new csv
        """
        with open(csv_path, 'wb') as f:
            c = csv.writer(f)
            for r in self.__sheet.rows:
                c.writerow([cell.value for cell in r])

    def get_cell(self, cell):
        """
        Get a cell
        :param cell: String. Cell coordinate
        :return: Cell.
        """
        return self.__sheet[cell]

    def get_cell_by_index(self, row, column):
        """
        Get a cell using coordinate.
        :param row: Integer. Number of row.
        :param column: Integer. Number of column.
        :return: Cell
        """
        cell = self.__sheet.cell(row, column)
        return cell

    def get_rows(self):
        """
        Get all of rows in the sheet.
        :return: Generator. Rows of active sheet.
        """
        min_r = self.get_min_row()
        max_r = self.get_max_col()
        return self.__sheet.iter_rows(min_row=min_r, max_row=max_r)

    def get_min_row(self):
        """
        Get first row
        :return: Integer. Number of first row
        """
        return self.__sheet.min_row

    def get_max_row(self):
        """
        Get first column.
        :return: Integer. Number of first column.
        """
        return self.__sheet.max_row

    def get_cols(self):
        """
        Get all of columns in the sheet
        :return: Generator. List of columns in the active sheet
        """
        min_c = self.get_min_col()
        max_c = self.get_max_col()
        return self.__sheet.iter_cols(min_col=min_c, max_col=max_c)

    def get_min_col(self):
        """
        Get the first column in the sheet.
        :return: Integer. Number of the first column.
        """
        return self.__sheet.min_column

    def get_max_col(self):
        """
        Get the last column in the sheet.
        :return: Integer. Number of the last column in the sheet.
        """
        return self.__sheet.max_column

    def __add_comment(self, cell, text, author):
        """
        Method to insert a comment in a cell.
        :param cell: Cell. Cell to insert a comment
        :param text: String. Text of comment
        :param author: String. Author of comment
        """
        comment = Comment(text, author)
        cell.comment = comment
        self.__save()

    def add_comment(self, cell, comment, author):
        """
        Insert a comment in a cell.
        :param cell: String. Cell to insert a comment
        :param comment: String. Text of the comment
        :param author: String. Author of the comment
        """
        cell = self.__sheet[cell]
        self.__add_comment(cell, comment, author)

    def add_comment_by_index(self, row, column, comment, author):
        """
        Insert a comment in a cell by index.
        :param row: Integer. Number of row.
        :param column: Integer. Number of column.
        :param comment: String. Text to insert in the cell.
        :param author: String. Author of the comment
        """
        cell = self.get_cell_by_index(row, column)
        self.__add_comment(cell, comment, author)

    def insert_row(self, row):
        """
        Insert a new row in the sheet.
        :param row: Integer. Position for the new row in the sheet.
        """
        self.__sheet.insert_rows(row)
        self.__save()

    def insert_column(self, column):
        """
        Insert a new column in the sheet.
        :param column: Integer. Position for the new column in the sheet
        """
        self.__sheet.insert_cols(column)
        self.__save()

    def move_row(self, row, newRow):
        """
        Move a row to a new position.
        :param row: Integer. Number of the row to move.
        :param newRow: Integer. Number where move the row.
        """
        cells = self.__sheet[row]
        self.delete_rows(row)
        self.insert_row(newRow)
        column = 0
        for cell in cells:
            paste = self.__sheet.cell(row=newRow, column=1)
            paste = paste.offset(column=column)
            self.__paste_cell(paste, cell)
            column += 1
        self.__save()

    def move_column(self, column, newColumn):
        """
        Move a column to new position
        :param column: Integer. Number of the column to move.
        :param newColumn: Integer. Number where move the column
        """
#         cols =self.get_cols()
#         counter = 1
#         while counter < column:
#             cols.next()
#             counter +=1
#         cells = cols.next()
        cells = self.get_column(column)
        self.delete_cols(column)
        self.insert_column(newColumn)
        row_aux = 0
        for cell in cells:
            paste = self.__sheet.cell(row=1, column=newColumn)
            paste = paste.offset(row=row_aux)
            self.__paste_cell(paste, cell)
            row_aux += 1
        self.__save()

    def change_column_size(self, column, width=None):
        """
        Method to change width of a column. If width is None, auto-resize.
        :param column: Integer. Number of the column to resize.
        :param width: Integer. New width for the column.
        """
        if width is None:
            col = self.__sheet[column]
            maximum = 0
            for cell in col:
                # Tengo q buscar la celda con mayor letra
                text = self.__read_cell_value(cell)
                if text is None:
                    pass
                else:
                    if len(text) > maximum:
                        maximum = len(text)
            width = maximum * 1.2 + 1
        self.__sheet.column_dimensions[column].width = width
        self.__save()

    def change_row_size(self, row, height=15):
        """
        Method to change height of a row.
        :param row: Integer. Number of row to change size.
        :param height: Integer. Height for column. Default value is 15
        """
        self.__sheet.row_dimensions[row].height = height
        self.__save()

    def __is_formula(self, cell):
        """
        If a cell is a formula
        :return: Boolean.
        """
        if python_version == 2:
            return cell.data_type == Cell.TYPE_FORMULA
        return cell.data_type == TYPE_FORMULA

    def search_value(self, value):
        """
        Search a value in all cells in the sheet.
        :param value: String. Searched value
        :return: list. List of cells coordinate.
        """
        min_r = self.get_min_row()
        max_r = self.get_max_row()
        aux = []
        for row in self.__sheet.iter_rows(min_row=min_r, max_row=max_r):
            for cell in row:
                if cell.value is None:
                    pass
                else:
                    if cell.value == value:
                        aux.append(cell.coordinate)
        return aux

    def replace_value(self, value, search_value):
        """
        Search and replace a value in all cells in the sheet.
        :param value: String. New value.
        :param search_value: String. Searched value.
        """
        aux = self.search_value(search_value)
        if aux:
            for coordinate in aux:
                self.write_cell_by_cell(coordinate, value)
            self.__save()

    def create_table(self, table_name, cell_range,
                     table_style=None):
        """
        Create a table using a range of cell.
        :param table_name: String. Name of the table.
        :param cell_range: String. Range of cell.
        :param table_style: dict. Dict of table format.
            example =dict(name="TableStyleMedium9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True,
                          showColumnStripes=True)
        """
        tab = Table(displayName=table_name, ref=cell_range)
        if table_style is not None:
            style = TableStyleInfo(
                name=table_style.get('name'),
                showFirstColumn=table_style.get('showFirstColumn'),
                showLastColumn=table_style.get('showLastColumn'),
                showRowStripes=table_style.get('showRowStripes'),
                showColumnStripes=table_style.get('showColumnStripes')
            )
            tab.tableStyleInfo = style
        self.__sheet.add_table(tab)
        self.__save()

    def __save(self, filename=None):
        """
        Method to save the excel file
        :param filename: String. Path of the excel file.
        """
        if filename is None:
            path = self.__path
        else:
            path = filename
        try:
            self.__xl.save(path)
        except IOError:
            exit()

    def save(self, filename):
        """
        Method to save the excel file
        :param filename: String. Path of the file.
        """
        self.__save(filename)

    def set_active_sheet(self, index=None, name=None):
        """
        Method to change the active sheet
        :param index: Integer. Number of the sheet to set active
        :param name: String. Name of sheet to set active
        """
        if index is not None:
            self.__xl.active = index
            self.__sheet = self.__xl.active
        elif name is not None:
            self.__sheet = self.__xl[name]
        else:
            print('Please enter the sheet\'s index or name to read')
        self.__head = self.read_row(1)


class usuarios(excel_file):
    def lectura_usuarios(self, columnas):
        max_r = self.get_max_row()
        contador = 2
        array = []
        while contador < max_r:
            fila = self.read_row(contador)
            aux = [contador]
            # Si no es una fila vacia, se procesa
            if fila[0] is not None:
                for columna in columnas:
                    aux.append(fila[columna])
                array.append(aux)
            contador += 1
        return array

    def __checkduplicate(self, array, fila):
        """
        Funcion para verificar que un usuario no se repita dentro de la matriz
        """
        # obtengo el usuario de la fila
        user = fila[1]
        for usuario in array:
            if user in usuario:
                return True
        return False

    def leer_usuario_fila(self, fila):
        """
        Obtengo los usuarios de una fila especifica
        """
        row = self.read_row(row=fila)
        return row

    def escritura_usuario(self, indice, columna, texto, color=C_WHITE):
        celda = self.get_cell_by_index(indice, columna)
        color = self.get_color(color)
        fill = self.set_cell_fill(color)
        celda.fill = fill
        self.write_cell_by_index(indice, columna, texto)

    def obtener_participante(self, fila):
        fila = self.leer_usuario_fila(fila)
        diccionario = {
            'docTipo': fila[6],
            'docNro': fila[2]
            }
        return diccionario

    def obtener_segundo_participante(self, fila):
        fila = self.leer_usuario_fila(fila)
        diccionario = {
            'docTipo': fila[6],
            'docNro': fila[2],
            'docTipo2': fila[7],
            'docNro2': fila[8]
            }
        return diccionario

    def buscar_caso(self, nombre_caso):
        """
        Metodo para buscar en la primera columna, la celda que tenga el nombre
        del caso, para poder realizar la lectura de los datos
        Devuelve el numero de la fila en la que se encuentra la celda
        encontrada, en caso que no se la encuentre, devuelve nulo.
        """
        columna = self.get_column(1)
        cell = None
        for celda in columna[1:]:
            if celda.value is not None:
                # casos = celda.value.replace(' ', '')
                # casos = str(casos)
                # casos = casos.split(',')
                # if nombre_caso in casos:
                #     cell = celda
                #     break
                if nombre_caso == celda.value:
                    cell = celda
                    break
        if cell is None:
            print('No se encontro el valor en la columna')
            return cell
        return cell.row
    
    def buscar_por_usuario(self, nombre_usuario):
        """
        Metodo para buscar en la primera columna, la celda que tenga el nombre
        del usuario, para poder realizar la lectura de los datos
        Devuelve el numero de la fila en la que se encuentra la celda
        encontrada, en caso que no se la encuentre, devuelve nulo.
        """
        columna = self.get_column(1)
        cell = None
        nombre_usuario = str(nombre_usuario)
        for celda in columna[1:]:
            if celda.value is not None:
                # casos = celda.value.replace(' ', '')
                # casos = str(casos)
                # casos = casos.split(',')
                # if nombre_caso in casos:
                #     cell = celda
                #     break
                if nombre_usuario == celda.value:
                    cell = celda
                    break
        if cell is None:
            print('No se encontro el valor en la columna')
            return cell
        return cell.row

    def obtener_datos_usuarios(self, nombre_caso, destino='hb'):
        """
        Metodo para obtener los datos autmaticamente desde el excel
        pasandole solo el nombre del caso
        :param nombre_caso: String. Nombre del caso
        :param destino: String. Lectura para hb o CRM
        """
        nombre_caso = str(nombre_caso)
        fila = self.buscar_caso(nombre_caso)
        if fila is not None:
            datos_test = self.read_row_by_head(fila)
            if destino == 'hb' or destino == 'Caja'or destino == 'mesaWeb' or destino == 'IBE' or destino =='BPM' or destino =='WhatsApp' or destino =='Genesys':
                usuario = datos_test.get('Usuario')
            else:
                return datos_test
            self.set_active_sheet(name='Usuarios')
            fila = self.__datos_usuario(usuario)
            datos_user = self.read_row_by_head(fila)
            datos_test.update(datos_user)
            # return dict(sorted(datos_test.items()))
            return datos_test
        return None
    
    def usuarios_activos(self, nombre_usuario, destino='hb'):
        
        nombre_usuario = str(nombre_usuario)
        fila = self.buscar_por_usuario(nombre_usuario)
        
        if fila is not None:
            datos_test = self.read_row_by_head(fila)
            if destino == 'hb' or destino == 'Caja'or destino == 'mesaWeb' or destino == 'IBE' or destino =='BPM':
                self.usuario = datos_test.get('Usuario')
                self.clave = datos_test.get('Clave')
            else:
                return datos_test
            return self.usuario, self.clave
        return None
            
        lista_activos = ["Usuarioautouno1", "Usuarioauto5", "Usuario95625072", 
                        "Usuarioauto2", "Usuarioauto4", "Usuarioautoe1", "Usuarioauto6", 
                        "Usuario32321585", "Usuario8672", "Usuario2561", "Usuario6987", "Usuario6308"]
        
        if nombre_usuario in lista_activos:
            excel=usuarios(hb_user_excel, hb_user_login)
            return excel.__datos_usuario(nombre_usuario)
        else:
            print("No se encuentra el usuario en la fila")
    
    def __datos_usuario(self, usuario):
        """
        Metodo para obtener los datos del usuario en la hoja de usuarios
        :param usuario: String. Nombre del usuario que se quiere leer
        """
        filas = self.get_column(1)
        index = None
        for fila in filas[1:]:
            if fila.value == usuario:
                index = fila
                break
        return index.row

    def obtener_usuarios(self):
        """
        Metodo para obtener todos los datos de los usuarios de la hoja de test
        """
        filas = self.get_column(1)
        datos = []
        for fila in filas[1:]:
            datos.append(self.read_row_by_head(fila.row))
        return datos

    def buscar_usuario(self, usuario, col=2):
        """
        Metodo para realizar la busqueda de un usuario en la segunda columna
        de la planilla excel.
        IMPORTANTE: Los usuarios tienen que estar en la segunda columna en la
        hoja de test. En la hoja de usuarios, estos se encuentran en la primera
        columna
        """
        columna = self.get_column(col)
        cell = None
        for celda in columna:
            if usuario in celda.value:
                cell = celda
                break
        if cell is None:
            print('No se encontro el valor en la columna')
            return cell
        return cell.row

    def actualizar_usuario(self, usuario, datos):
        """Metodo para actualizar los datos del usuario"""
        self.set_active_sheet(1)
        fila = self.buscar_usuario(str(usuario), col=1)
        self.write_row_by_head(fila, datos)

